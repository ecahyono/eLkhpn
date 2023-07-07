#Untuk kebutuhan data Penerimaan Rupbasan
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from pytest_html_reporter import attach
from pytest import mark
import pyautogui
import platform
import logging
import subprocess
import os, platform, time, pytest
from os import environ, path
from pathlib import Path
import sys
from openpyxl import load_workbook
from dotenv import load_dotenv
from faker import Faker
from faker.providers import date_time
from datetime import datetime, timedelta
import random

fake = Faker()
load_dotenv()

Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
fh = logging.FileHandler('result.log', mode="a")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

from setupbrowser import initDriver, loadDataPath
from userlogin import *

if platform.system() == 'Darwin':
    wb = load_workbook(environ.get("dataexel"))
elif platform.system() == 'Windows':
    wb = load_workbook(environ.get("dataexel"))


################-> input urutan Baris Di Exel yang akan di eksekusi
pendampingan = 6
litmas = 5
pembimbingan = 3
pengawsan = 12
curd = 2

sheet = wb['PKpendampingan']
opsiJenisLayananBerikutnya    = [cell.value for cell in sheet['A'] if cell.value is not None]

fake = Faker('id_ID')

############################## PK->Pendampingan
fake.add_provider(date_time)
TanggalPendampingan           = fake.date_time().strftime('%d/%m/%Y %H:%M')
NamaPenyidik                  = fake.first_name()
ResumeBeritaAcaraPendampingan = fake.random_element(elements=('Berita acara Pendampingan Pertama', 'Berita acara Pendampingan Kedua'))
TanggalPendampinganBerikutnya = datetime.strptime(fake.date(), '%Y-%m-%d').strftime('%d/%m/%Y')
JenisLayananBerikutnya        = fake.random.choice(opsiJenisLayananBerikutnya)
Switchfield                   = fake.random_element(elements=('Iya', 'Tidak'))
TipeLampiran                  = fake.random_element(elements=('Pendampingan Diversi', 'Pendampingan Pelaksanaan Kesepakatan Diversi'))
TingkatPengadilan             = fake.random_element(elements=('Negeri', 'Tinggi','Mahkamah'))

tanggal_sekarang = datetime.now()
tanggal_sebelumnya = tanggal_sekarang - timedelta(days=30)
tgl_sebulan_berikutnya = tanggal_sekarang + timedelta(days=30)

A = wb['Register Pembimbingan']
UPTObimbingan                           = A['A'+str(pembimbingan)].value
NoregNamabimbingan                      = A['B'+str(pembimbingan)].value
jenisKlienPembimbingan                  = A['C'+str(pembimbingan)].value
JenisPembimbingan                       = A['D'+str(pembimbingan)].value
DasarPembimbingan                       = A['E'+str(pembimbingan)].value
TglAwalBimbingan                        = tanggal_sebelumnya.strftime('%d/%m/%Y')
TglAkhirBimbignan                       = tgl_sebulan_berikutnya.strftime('%d/%m/%Y')
CariPetugasPembimbingan                 = A['H'+str(pembimbingan)].value
SuratDasarPembimbingan                  = A['I'+str(pembimbingan)].value
AsalsuratPembimbingan                   = A['J'+str(pembimbingan)].value
Nosurat1Pembimbingan                    = A['K'+str(pembimbingan)].value
tglsuratPembimbingan                    = tanggal_sekarang.strftime('%d/%m/%Y %H:%M')   
Perihalsurat1Pembimbingan               = A['M'+str(pembimbingan)].value   
SuratPengantarPenyerahanPembimbingan    = A['N'+str(pembimbingan)].value   
BeritaAcaraSerahTerimaKlienPembimbingan = A['R'+str(pembimbingan)].value
SuratPerintahPembimbingan               = A['V'+str(pembimbingan)].value