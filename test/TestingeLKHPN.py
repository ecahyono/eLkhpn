from selenium import webdriver 
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from pytest_html_reporter import attach
import os, platform, time, pytest
from openpyxl import Workbook
from os import environ, path
from pathlib import Path
from faker import Faker
from pytest import mark
import subprocess
import pyautogui
import platform
import logging
import sys
from openpyxl import load_workbook
from dotenv import load_dotenv
load_dotenv()

if platform.system() == 'Darwin':
    sys.path.append(environ.get("MACPARENTDIR")) 
elif platform.system() == 'Windows':
    sys.path.append(environ.get("WINPARENTDIR"))

from config.setwebdriver import initDriver, loadDataPath, secondaryinit, webfirefox
from config.userlogin import *


Log = logging.getLogger(__name__)
log_format = '[%(asctime)s %(filename)s->%(funcName)s()]==>%(levelname)s: %(message)s'
timestamp = time.strftime("%Y%m%d_%H%M%S", time.localtime())
fh = logging.FileHandler(f'registrasi_{timestamp}.log', mode="a")
fh.setLevel(logging.INFO)
formatter = logging.Formatter(log_format)
fh.setFormatter(formatter)
Log.addHandler(fh)

#fileexcel
if platform.system() == 'Darwin':
    wb = load_workbook(environ.get("excelsatker"))
elif platform.system() == 'Windows':
    wb = load_workbook(environ.get("dataexel"))

# init driver by os
def testconfigandlogin():
	Log.info('Konfigurasi agar berjalan di setiap sistem operasi (mac dan Windos)')
	global driver, pathData
	driver = initDriver()
	pathData = loadDataPath()
	Log.info('Memasukan User name dan Password di halaman Login')





